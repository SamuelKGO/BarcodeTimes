import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import os

def export_to_excel(data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Barcode", "Start Time", "End Time", "Duration (minutes)"])

    # adds headers for excel sheet and determines date and time for file name
    for barcode, start_time, end_time, duration in data:
        sheet.append([barcode, start_time, end_time, duration])
    date_today = datetime.datetime.today().strftime('%m-%d-%Y -- %I %M %S %p')
    file_name = "TS_data_" + date_today + ".xlsx"


    column_widths = []
    # adjusts cell widths to match text size #############################
    for row in data:
        print(list(enumerate(row)))
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell)) > column_widths[i]:
                        column_widths[i] = len(str(cell))
            else:
                print(cell)
                column_widths += [len(str(cell)) + 1]

    for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
        sheet.column_dimensions[get_column_letter(i)].width = column_width
    #######################################################################


    def UI(file_name, date_today, workbook):
        answer = input("Export to default directory? [y\\n]: ").lower()
        if answer in ['n', 'no']:
            # successful path for custom directory
            try:
                new_dir = input("Enter desired directory destination: ")
                os.chdir(new_dir)
                workbook.save(file_name)
                return new_dir
            # error correction for custom directory (recursive call)
            except FileNotFoundError:
                print("\n\nDIRECTORY NOT FOUND PLEASE RETRY\n\n ")
                UI(file_name, date_today, workbook)
        # succesful path for default directory
        elif answer in ['y', 'yes']:
            new_dir = 'O:\\SHEET METAL\\0 -- Samuel Ogura (intern)\\Barcode-Time Studies' # Replace with your desired default directory path
            os.chdir(new_dir)
            workbook.save(file_name)
            return new_dir
        # error for ui input (recursive call)
        else:
            print("\n\nPLEASE ENTER Y/YES OR N/NO\n\n")
            UI(file_name, date_today, workbook)

    return file_name, UI(file_name, date_today, workbook)
def main():
    def funct(barcode):
        keystrokes = []
        count = []
        time_records = {}
        data = []
        counter = 0

        while barcode not in ["Q", "q"]:  # terminate loop when 'q' is entered
            duration = track_time(count, time_records, barcode)
            count.append(barcode)
            time_records[barcode] = datetime.datetime.now()
            print(f"BARCODE:\t{barcode}\nTIME:  \t{time_records[barcode]}\n")

            if duration is not None:
                start_time = time_records[barcode]
                end_time = start_time + datetime.timedelta(minutes=duration)
                data.append([barcode, str(start_time), str(end_time), duration])
                print(f'\ndata: {data[counter]}\n\n')
                counter += 1

            barcode = input("Scan: ")  # prompt for next barcode input

        name_and_dir = export_to_excel(data)
        print(f"\n\n\tData exported as:\t{name_and_dir[0]}\n\n\tto directory:\t\t{name_and_dir[1]}")

    barcode = input("Scan: ")
    funct(barcode)

def track_time(count, time_records, barcode):
    if barcode in count:
        start_time = time_records[barcode]
        end_time = datetime.datetime.now()
        time_difference = end_time - start_time
        del time_records[barcode]
        return ((time_difference.total_seconds())/60)
    else:
        return None

if __name__ == '__main__':
    main()
