import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import os
import logging
import time


def export_to_excel(data, station):

    # creates excel file to dump data in and creates headers
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    headers = ["Barcode Value", "Start Time", "End Time", "Duration (minutes)"]
    sheet.append(headers)

    # center formats  and emboldens headers
    for header_column in range(len(headers)):
        sheet.cell(row = 1,column = header_column + 1).alignment = openpyxl.styles.Alignment(horizontal = 'center')
        sheet.cell(row=1, column=header_column + 1).font = openpyxl.styles.Font(bold=True)

    # adds title for excel file using station, date and time
    for barcode, start_time, end_time, duration in data:
        sheet.append([barcode, start_time, end_time, (str(duration) + " min")])
    date_today = datetime.datetime.today().strftime('%m-%d-%Y -- %I %M %S %p')
    file_name = station + " TS_data_" + date_today + ".xlsx"

    # create list to store column widths
    column_widths = []

    # set cells to minimum of header widths
    for title in headers:
        column_widths += [len(str(title)) + 1]

    # adjusts cell widths to match text size #############################
    for row in data:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell)) > column_widths[i]:
                        column_widths[i] = (len(str(cell)) + 1)
            else:
                column_widths += [len(str(cell)) + 1]

    for i, column_width in enumerate(column_widths, 1):  # ,1 to start at 1
        sheet.column_dimensions[get_column_letter(i)].width = column_width
    #######################################################################

    # creating and populating merged cell to store copy/paste AX lookup data
    sheet.merge_cells(start_row = 2, start_column = 6, end_row = 6, end_column = 8)
    sheet.cell(row = 4, column = 9).value = "<------  COPY/PASTE INTO AX"
    sheet.cell(row = 4, column = 9).font = openpyxl.styles.Font(bold=True)
    highlight = openpyxl.styles.fills.PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid") #adjust rgb values to whatever color
    sheet.cell(row = 2, column = 6).fill = highlight

    copypasta = []
    for barcode_val in data:
        raw_val = '*' + str(barcode_val[0]) + '*'
        copypasta.append(raw_val)
    sheet.cell(row = 2, column = 6).value = str(set(copypasta)).strip("{}").replace("'","")


    def UI(file_name, date_today, workbook):
        answer = input("\tExport to default directory? [y/n]: ").strip().lower()
        if answer in ['n', 'no']:
            # successful path for custom directory
            try:
                new_dir = input("\tEnter desired directory destination: ").strip()
                os.chdir(new_dir)
                workbook.save(file_name)
            # error correction for custom directory (recursive call)
            except FileNotFoundError:
                print("\n\n\tDIRECTORY NOT FOUND PLEASE RETRY\n\n ")
                UI(file_name, date_today, workbook)
        # succesful path for default directory
        elif answer in ['y', 'yes']:
            new_dir = 'O:\\SHEET METAL\\0 -- Barcode TimeStudies' # replace with your desired default directory path
            os.chdir(new_dir)
            workbook.save(file_name)
        # error for ui input (recursive call)
        else:
            print("\n\n\tPLEASE ENTER Y/YES OR N/NO\n\n")
            UI(file_name, date_today, workbook)
        return new_dir
    return file_name, UI(file_name, date_today, workbook)



def track_time(state, time_records, barcode):

    if barcode in state and state[barcode] == "in":

        start_time = time_records[barcode]
        end_time = datetime.datetime.now()
        time_difference = end_time - start_time
        print(f"\n\n\tthis is the data you are looking for: {time_records[barcode]}")
        del time_records[barcode]
        return (round((time_difference.total_seconds()/60),2))
    else:
        return None



def main():
    def funct(barcode):
        keystrokes = []
        state = {}
        time_records = {}
        data = []
        counter = 0

        while barcode not in [ "q", "Q", "Quit", "quit", "QUIT", "EXIT", "Exit", "exit"]:  # terminate loop when 'q' is entered
            duration = track_time(state, time_records, barcode)
            state[barcode] = "in"
            time_records[barcode] = datetime.datetime.now()
            print(f"\tBARCODE:\t{barcode}\n\tTIME:  \t{time_records[barcode]}\n")
            if duration is not None:
                state[barcode] = "out"
                start_time = time_records[barcode]
                end_time = start_time + datetime.timedelta(minutes=duration)
                data.append([barcode, str(start_time), str(end_time), duration])
                print(f'\n\tDATA  -->  BARCODE VALUE: {data[counter][0]} | BEGIN TIME: {data[counter][1]} | END TIME: {data[counter][2]} | ELAPSED TIME: {data[counter][3]} minutes\n\n')
                counter += 1
            barcode = input("\n\t(q, quit, or exit to end)\n\tScan: ").strip()  # prompt for next barcode input

        name_and_dir = export_to_excel(data, station)
        print(f"\n\n\tData exported as:\t{name_and_dir[0]}\n\n\tto directory:\t\t{name_and_dir[1]}")

    def restart():
        # recursively call main to continue making more logs
        decision = input("\n\n\n\tWant to log another file? [y/n]:").strip().lower()
        if decision in ("yes", "y", "no", "n"):

            while decision in ("yes", "y"):
                main()
            time.sleep(1)
            print("\n\n\n\n     Closing Program...")
            time.sleep(3)

        else:
            print("\n\tERROR, enter yes or no")
            restart()

    station = input("\n\twhat station is this at?:").strip()
    barcode = input("\n\t(q, quit, or exit to end)\n\tScan: ").strip()
    funct(barcode)
    restart()


if __name__ == '__main__':

    try:
        # set up logging configuration at beginning of script
        log_directory = "O:\\SHEET METAL\\0 -- Barcode TimeStudies\\ERROR LOGS"
        if not os.path.exists(log_directory):
            os.makedirs(log_directory)

        # call the main function
        main()

    except Exception as e:
        print('\n\n\tERROR ERROR ERROR ERROR\n\n')
        log_name = "error_log_0.txt"
        while os.path.exists(log_directory + "\\" + log_name):
            log_num = int("".join(filter(str.isdigit, log_name)))
            stripped_name = log_name.strip("0123456789.txt")
            log_num += 1
            log_name = stripped_name + str(log_num)
            log_name += ".txt"
        log_file_path = os.path.join(log_directory, log_name)
        log_directory += "\\"
        print(f"\n\tEXPORTING ERROR LOG AS: {log_name}\n\tPATH: {log_directory + log_name}")
        logging.basicConfig(filename=log_file_path, level=logging.ERROR)

        # log an error with traceback
        logging.exception("An error occurred:")
        print(f"\tthis error occurred: {str(e)}")
        input("\n\tPress Enter to exit...")

    #updated code
